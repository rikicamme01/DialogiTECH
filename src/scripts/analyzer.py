import pandas as pd
import sys
import os
import itertools
sys.path.append(os.path.dirname(sys.path[0]))
from ast import literal_eval
from datasets.ie_hyperion_dataset import find_word_bounds, clean_text
from models.bert_segmenter import BertSegmenter
from models.bert_rep import BertRep
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datasets.hyperion_dataset import LABELS 

class analyzer():
    def __init__(self) -> None:
        bert_seg = BertSegmenter()
        bert_rep = BertRep()
    
    
    def predict(self, path_file):
        extension = os.path.splitext(path_file)[1]

        if extension == '.xlsx':
            df = pd.read_excel(path_file, converters={'Stralci': literal_eval, 'Repertori': literal_eval} )
        elif extension == '.csv':
            df = pd.read_csv(path_file, converters={'Stralci': literal_eval, 'Repertori': literal_eval})
        else:
            raise Exception('Extension of file not supported')
        
        #carica csv
        if 'Testo' in df.columns:
            df['Testo'] = df['Testo'].map(clean_text)
            #df['Stralci'] = df['Stralci'].map(lambda x: [clean_text(s) for s in x])
            #df['Bounds'] = df.apply(lambda x: find_word_bounds(x['Stralci'], x['Testo']), axis=1).values.tolist()
        else:
            raise Exception('The uploaded file is missing the "Testo" column')

        #predict
        df['Stralci_predetti'] = df['Testo'].map(self.bert_seg.predict).values.tolist()
        #df['Bounds_predetti'] = df.apply(lambda x: find_word_bounds(x['Stralci_predetti'], x['Testo']), axis=1).values.tolist()
        df['Repertori_predetti'] = df['Stralci_predetti'].map(self.bert_rep.predict_vector).values.tolist()
        #df['Repertori_predetti'] = df['Stralci_predetti'].apply(bert_rep.predict_vector)
   
        self.df_to_excel(df)



    def df_to_excel(df):
        wb = Workbook()
        ws = wb.active
        columns_format(ws)

        for row in df.iterrows():
            stralci_csv = row['Stralci_predetti'] # lista di stralci per quel testo
            repertori_csv = row['Repertori_predetti'] # lista di dizionari per ciascuno stralcio

            if row['Testo']:
                for index, stralcio in stralci_csv:
                    dict_current = repertori_csv[index] # dizionario dello stralcio attuale
                    if index == 0:
                        ws['Testo'].append(row['Testo'])
                    else:
                        ws['Testo'].append('')
                    ws['Stralcio'].append(stralcio)
                    # Aggiungi il nuovo dizionario alla lista dei nuovi dati
                    for rep in LABELS:
                        ws[f'{rep}'].append(dict_current[rep])
                    # cerco repertorio con la probabilità maggiore
                    max_prob = max(dict_current.values())
                    for key, value in dict_current:
                        if value == max_prob:
                            max_rep = key 
                    ws['Repertorio predetto'].append(max_rep)
                    ws['Accuratezza'].append(max_prob)
                    
        #scorro tutta la tabella e controllo i valori di 'Accuratezza' e nel caso sottolineo
        for row in ws.iter_row(values_only =True):
            if row['Accuratezza'] < 0.5:
                pass

        wb.save('prova.xlsx')

    
    def columns_format(ws):
        #creazione colonne intestazione
        # set up DataValidation
        columns = ['Domanda', 'Età', 'Genere', 'Ruolo', 'Testo', 'Stralcio', 'Repertorio predetto','Accuratezza/proposte']

        validation = DataValidation(type = 'list', formula1 = lambda x: ','.join(LABELS))
        ws.add_data_validation(validation)
        validation.add(ws['Repertorio predetto'])
